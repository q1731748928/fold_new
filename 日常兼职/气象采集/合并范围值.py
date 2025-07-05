from openpyxl import load_workbook

# 加载 Excel 文件
wb = load_workbook('/Users/hejie/PycharmProjects/turing_exec/日常兼职/气象采集/qxcj_final.xlsx')
sheet = wb.active  # 获取当前活动的工作表

# 假设数据从第二行开始，第1行是表头
for row in range(2, sheet.max_row + 1):
    # 获取E列和F列的值
    e_value = str(sheet[f'E{row}'].value)
    f_value = str(sheet[f'F{row}'].value)

    # 拼接E列和F列的值，赋值给D列
    sheet[f'D{row}'] = e_value + "~" + f_value

# 保存修改后的文件
wb.save('qxcj_final.xlsx')
