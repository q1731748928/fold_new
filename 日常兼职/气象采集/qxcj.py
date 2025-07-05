import time

import pymysql
from openpyxl.reader.excel import load_workbook

db = pymysql.connect(host="localhost", port=3306, user="root", passwd="123456", db="qxcj")
# 使用 cursor() 方法创建一个游标对象 cursor
cursor = db.cursor()

import execjs
import requests

# 要用自己的cookie
cookies = {
    'Hm_lvt_6088e7f72f5a363447d4bafe03026db8': '1704421259',
    'Hm_lpvt_6088e7f72f5a363447d4bafe03026db8': '1704422591',
}

headers = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Connection': 'keep-alive',
    'Referer': 'https://www.aqistudy.cn/historydata/',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'same-origin',
    'Sec-Fetch-User': '?1',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'sec-ch-ua': '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
}

# 要查什么城市自己改这个参数即可
# 读取Excel文件
wb = load_workbook('/Users/hejie/PycharmProjects/turing_exec/日常兼职/气象采集/qxcj.xlsx')  # 替换为你的文件路径
sheet = wb.active  # 获取活动的工作表

# 创建一个新的容器来存储处理后的值
new_container = []

# 获取第2列的所有数据并处理
for row in range(2, sheet.max_row + 1):  # 遍历每一行
    value = sheet.cell(row=row, column=2).value  # 获取第2列的值

    # 如果该值是字符串且包含“、”，将其拆分成多个字符串
    if isinstance(value, str) and '、' in value:
        split_values = value.split('、')  # 按“、”拆分字符串
        new_container.extend(split_values)  # 将拆分后的值加入容器
    elif value is not None:
        new_container.append(value)  # 如果不是字符串，直接添加到容器

# 遍历输出容器中的所有值
for item in new_container:
    city = str(item)
    # 读取js代码，构造js执行环境
    with open("qxcj.js", 'r', encoding='utf-8') as f:
        js_code = f.read()
    env = execjs.compile(js_code)
    query = env.call("gethistory", city)
    data = {
        'hA4Nse2cT': query
    }
    response = requests.post('https://www.aqistudy.cn/historydata/api/historyapi.php', cookies=cookies, headers=headers,
                             data=data)
    time.sleep(30)
    encrypted_data = response.text
    result = (env.call("get_data", encrypted_data))
    items = result['result']['data']['items']
    for item in items:
        time_point = item.get('time_point')
        aqi = item.get('aqi')
        max_aqi = item.get('max_aqi')
        min_aqi = item.get('min_aqi')
        pm2_5 = item.get('pm2_5')
        pm10 = item.get('pm10')
        co = item.get('co')
        no2 = item.get('no2')
        o3 = item.get('o3')
        so2 = item.get('so2')
        so2 = item.get('so2')
        rank = item.get('rank')
        quality = item.get('quality')
        range = f'{min_aqi}~{max_aqi}'
        print(f"{city} {time_point} {aqi} {max_aqi} {min_aqi} {range} {pm2_5} {pm10} {co} {no2} {o3} {so2} {rank} {quality}")
        # try:
        #     sql = "INSERT INTO qxcj (city, time_point, aqi, max_aqi, min_aqi, pm2_5, pm10, co, no2, o3, so2, quality) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s,%s)"
        #     val = (city, time_point, aqi, max_aqi, min_aqi, pm2_5, pm10, co, no2, o3, so2, quality)
        #     cursor.execute(sql, val)
        #     db.commit()
        # except Exception as e:
        #     print(e.args)
