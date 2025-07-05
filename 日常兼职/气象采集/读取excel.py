from openpyxl import load_workbook

# 读取Excel文件
wb = load_workbook('/Users/hejie/PycharmProjects/turing_exec/日常兼职/气象采集/qxcj.xlsx')  # 替换为你的文件路径
sheet = wb.active  # 获取活动的工作表

# 创建一个新的容器来存储处理后的值
new_container = []

# 获取第2列的所有数据并处理
for row in range(2, sheet.max_row + 1):  # 遍历每一行
    value = sheet.cell(row=row, column=2).value  # 获取第2列的值

    # 如果该值是字符串且包含“、”，将其拆分成多个字符串
    if isinstance(value, str) and '、' in value:
        split_values = value.split('、')  # 按“、”拆分字符串
        new_container.extend(split_values)  # 将拆分后的值加入容器
    elif value is not None:
        new_container.append(value)  # 如果不是字符串，直接添加到容器
print(len(new_container))
# 遍历输出容器中的所有值
for item in new_container:
    print(item)