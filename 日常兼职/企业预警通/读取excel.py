import openpyxl
import pymysql
from openpyxl.reader.excel import load_workbook

from openpyxl import load_workbook

# 加载Excel文件
wb = load_workbook('/Users/hejie/PycharmProjects/turing_exec/日常兼职/企业预警通/2024.06在建工程.xlsx')  # 请替换为你的文件名
db = pymysql.connect(host="localhost", port=3306, user="root", passwd="123456", db="qyyjt")
# 使用 cursor() 方法创建一个游标对象 cursor
cursor = db.cursor()

# 获取活动的工作表
sheet = wb.active

# 用来存储第2列不重复的值
unique_values = set()


def find_exist(comment):
    # 假设我们要检查的用户名

    # SQL 查询语句，检查 username 是否存在
    query = "SELECT COUNT(*) FROM qyyjt WHERE company_name = %s;"

    # 执行查询
    cursor.execute(query, (comment,))

    # 获取查询结果
    result = cursor.fetchone()

    # 如果查询结果大于0，说明该用户名已经存在
    return result[0]


# 遍历第6行到第629行，获取第2列的数据
for row in sheet.iter_rows(min_row=6, max_row=629, values_only=True):
    unique_values.add(row[1])  # row[1] 表示第2列

# 打印不重复的数据
print("不重复的数据：")
for value in unique_values:
    result = find_exist(value)
    if result == 0:
        print(value)
# 打印不重复数据的数量
print(f"\n不重复的数据总数：{len(unique_values)}")


"""
中钢设备有限公司
宝钢湛江钢铁有限公司
宝钢工程技术集团有限公司
神雾环保技术股份有限公司
马鞍山钢铁股份有限公司
中国重型机械研究院股份公司
宝钢德盛不锈钢有限公司
上海宝成钢结构建筑有限公司
宝钢集团新疆八一钢铁有限公司
山东泰山钢铁集团有限公司
山东传洋集团有限公司
四川绿建杭萧钢构有限公司
扬州恒润海洋重工有限公司
山东钢铁集团日照有限公司
山东省冶金设计院股份有限公司
中天钢铁集团有限公司

"""