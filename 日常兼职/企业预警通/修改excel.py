import openpyxl
import pymysql

# 连接到 MySQL 数据库
conn = pymysql.connect(host='localhost', user='root', password='123456', db='qyyjt')

# 创建游标对象
cursor = conn.cursor()


def query(company_name):
    # 查询特定字段
    cursor.execute("SELECT main_rating,all_creditors_right,continuous_creditors_right FROM qyyjt WHERE company_name = %s", (company_name,))
    # 获取查询结果
    result = cursor.fetchone()
    if result:
        print(f"查询结果: main_rating: {result[0]}, all_creditors_right: {result[1]}, continuous_creditors_right: {result[2]}")
        return result
    else:
        print("未找到匹配的记录。")


# 打开文件
file_path = '/Users/hejie/PycharmProjects/turing_exec/日常兼职/企业预警通/2024.06在建工程.xlsx'  # 请替换为你的文件路径
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active  # 默认读取第一个工作表
last_column = sheet.max_column
# 打开 Excel 文件
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active  # 默认读取第一个工作表

# 设置你要修改的内容
new_value = "Updated Value"

# 获取工作表的最大列数
max_col = sheet.max_column

# 获取最后三列的列字母
last_col = chr(64 + max_col)  # 最后一列的列字母
second_last_col = chr(64 + max_col - 1)  # 倒数第二列的列字母
third_last_col = chr(64 + max_col - 2)  # 倒数第三列的列字母

# 初始化保存当前行第二列的值的列表
second_column_values = []

for row in range(14, 1154):
    # 获取当前行第二列的值并保存
    value_at_row_2 = sheet[f'B{row}'].value  # 获取当前行的第二列值
    print(f'行号:{row},当前所在公司是:{value_at_row_2}')
    # 检查该值是否包含 "政府", "大学", "中心", "局" 中的任意一个关键词
    if any(keyword in str(value_at_row_2) for keyword in ["政府", "大学", "中心", "局"]):
        print(f"行号:{row}，当前所在公司包含关键词，准备删除该行。")
        sheet.delete_rows(row)  # 删除当前行
    else:
        # 检查该值是否不等于 '——'
        if str(value_at_row_2) != '——':
            result = query(value_at_row_2)
            if result is not None and result[0]:
                # 修改当前行的最后三列的值
                sheet[f'M{row}'] = result[0]
                sheet[f'N{row}'] = result[1]
                sheet[f'O{row}'] = result[2]

# # 打印第二列的所有值
# print(second_column_values)
#
# # 定位到第 14 行
# target_row = 14
#
# # 获取当前工作表的最后一列
# last_column = sheet.max_column
#
# # # 添加字段到最后三列
# # sheet.cell(row=target_row, column=last_column + 1, value="企业的评级")
# # sheet.cell(row=target_row, column=last_column + 2, value="全部债权")
# # sheet.cell(row=target_row, column=last_column + 3, value="存续债权")
#
# # 更新后的新数据列表
# new_values = ['新数据1', '新数据2', '新数据3']
#
# # 修改这三列的值
# sheet.cell(row=target_row, column=last_column + 1, value=new_values[0])
# sheet.cell(row=target_row, column=last_column + 2, value=new_values[1])
# sheet.cell(row=target_row, column=last_column + 3, value=new_values[2])
#
# 保存更改后的 Excel 文件
workbook.save('your_file_updated.xlsx')  # 保存为新的文件，或者覆盖原文件

print("字段已成功添加!")
