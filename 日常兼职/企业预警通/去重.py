from collections import defaultdict

import openpyxl
from datetime import datetime

# 载入工作簿并选择活动表
wb = openpyxl.load_workbook('/Users/hejie/PycharmProjects/turing_exec/日常兼职/企业预警通/your_file_updated.xlsx')  # 替换成你的文件路径
sheet = wb.active

# 创建一个字典，用于存储每个公司最新的日期及其对应的行号
latest_rows = {}

def parse_date(date_str):
    # 输入日期字符串
    # 将字符串转换为 datetime 对象
    date_obj = datetime.strptime(date_str, '%Y-%m-%d')
    # 将 datetime 对象格式化为带零填充的字符串
    formatted_date = date_obj.strftime('%Y-%m-%d')
    return formatted_date
# 创建一个字典来存储每个公司对应的所有行和日期
company_rows = defaultdict(list)
# 遍历指定范围的行，从第14行开始到第1453行
for row in range(14, 1153):  # 1454是因为range是上限排除的
    company_name = sheet.cell(row=row, column=2).value  # 公司名称（假设在B列，即第2列）
    date_value = sheet.cell(row=row, column=12).value  # 日期值（假设日期在L列，即第12列）
    # 将公司和对应的行号及日期存储到字典中
    if date_value is not None:  # 仅当日期不为空时才存储
        company_rows[company_name].append((row, date_value))
# 对每个公司，按日期排序，保留日期较大的，删除日期较小的
for company_name, rows in company_rows.items():
    # 按日期降序排序，如果日期相同则保留第一条
    rows.sort(key=lambda x: x[1], reverse=True)
    # 用于追踪哪些行需要删除
    seen_dates = set()
    rows_to_delete = []
    # 遍历并标记需要删除的行
    for row, date_value in rows:
        if date_value not in seen_dates:
            seen_dates.add(date_value)  # 该日期首次出现，保留
        else:
            rows_to_delete.append(row)  # 日期已存在，标记删除该行
    # 从最后一行开始删除，以避免行索引变化
    for row in sorted(rows_to_delete, reverse=True):
        sheet.delete_rows(row)

# 保存修改后的Excel文件
wb.save('qxcj_final.xlsx')  # 请替换为你想保存的文件名