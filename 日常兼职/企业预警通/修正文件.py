from openpyxl import load_workbook

# 加载Excel文件
wb = load_workbook('/Users/hejie/PycharmProjects/turing_exec/日常兼职/企业预警通/2024.06在建工程.xlsx')  # 请替换为你的文件名

# 获取活动的工作表
sheet = wb.active

# 要删除的关键字
keywords = ['政府', '大学', '中心', '局']

# 收集需要删除的行索引
rows_to_delete = []

# 遍历第6行到第629行的数据
for row in range(5, 630):  # 从第6行到第629行
    cell_value = sheet.cell(row=row, column=2).value  # 获取第2列的数据
    if any(keyword in str(cell_value) for keyword in keywords):  # 检查是否包含关键字
        rows_to_delete.append(row)

# 删除行（从后往前删除，避免影响行号）
for row in reversed(rows_to_delete):
    sheet.delete_rows(row)

# 保存修改后的Excel文件
wb.save('qxcj_final.xlsx')  # 请替换为你想保存的新文件名
