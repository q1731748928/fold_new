import openpyxl
from openpyxl import Workbook
from datetime import datetime

def filter_excel(file_path):
    # 加载工作簿
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 假设公司名称在第一列，日期在第 12 列（L 列）
    company_col = 1
    date_col = 12

    # 存储每个公司的最大日期及其对应行数据
    company_data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
        company = row[company_col - 1]
        date_str = row[date_col - 1]
        if isinstance(date_str, str):
            try:
                date = datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                continue
        elif isinstance(date_str, datetime):
            date = date_str
        else:
            continue

        if company not in company_data:
            company_data[company] = (date, row)
        else:
            if date > company_data[company][0]:
                company_data[company] = (date, row)

    # 创建新的工作簿
    new_wb = Workbook()
    new_ws = new_wb.active

    # 复制表头
    header = [cell.value for cell in ws[1]]
    new_ws.append(header)

    # 将筛选后的数据添加到新工作簿
    for _, (_, row) in company_data.items():
        new_ws.append(row)

    return new_wb

if __name__ == "__main__":
    # 替换为你的 Excel 文件路径
    file_path = '/Users/hejie/PycharmProjects/turing_exec/日常兼职/企业预警通/your_file_updated.xlsx'
    result_wb = filter_excel(file_path)

    # 保存结果到新的 Excel 文件
    result_wb.save('filtered_excel_openpyxl.xlsx')