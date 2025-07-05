import openpyxl
from datetime import datetime

def filter_excel(input_file_path, output_file_path):
    # 加载工作簿
    wb = openpyxl.load_workbook(input_file_path)
    ws = wb.active

    # 存储每个公司对应的最大日期行索引
    company_max_date = {}

    # 遍历所有行，从第二行开始（跳过表头）
    for row_idx in range(2, ws.max_row + 1):
        company = ws.cell(row=row_idx, column=2).value
        date_cell = ws.cell(row=row_idx, column=12).value

        try:
            if isinstance(date_cell, str):
                date = datetime.strptime(date_cell, '%Y-%m-%d')
            elif isinstance(date_cell, datetime):
                date = date_cell
            else:
                continue

            if company not in company_max_date:
                company_max_date[company] = (date, row_idx)
            else:
                if date > company_max_date[company][0]:
                    company_max_date[company] = (date, row_idx)
        except ValueError:
            continue

    # 找出需要保留的行索引，包括表头
    rows_to_keep = [1] + [row_num for _, row_num in company_max_date.values()]

    # 创建新的工作簿
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active

    # 复制表头
    for col_idx in range(1, ws.max_column + 1):
        new_ws.cell(row=1, column=col_idx, value=ws.cell(row=1, column=col_idx).value)

    # 将筛选后的行添加到新工作簿
    new_row_idx = 2
    for row_num in rows_to_keep[1:]:
        for col_idx in range(1, ws.max_column + 1):
            new_ws.cell(row=new_row_idx, column=col_idx, value=ws.cell(row=row_num, column=col_idx).value)
        new_row_idx += 1

    # 保存结果到新文件
    new_wb.save(output_file_path)

if __name__ == "__main__":
    # 原文件路径
    input_file_path = '/Users/hejie/PycharmProjects/turing_exec/日常兼职/企业预警通/your_file_updated.xlsx'
    # 新文件路径，可根据需求修改
    output_file_path = '/Users/hejie/PycharmProjects/turing_exec/日常兼职/企业预警通/update.xlsx'
    filter_excel(input_file_path, output_file_path)